"""
generate_truth_data.py
======================
One-time script: writes projects/UM_Dearborn/truth_data.xlsx from the
AADT and hourly distribution data currently hard-coded in truth_data.py.

Run once from the repo root:
    python generate_truth_data.py
"""

from pathlib import Path
import numpy as np
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment

PROJECT_DIR = Path("projects/UM_Dearborn")

HEADER_FONT = Font(bold=True, color="FFFFFF")
HEADER_FILL = PatternFill("solid", fgColor="2E4057")

def _header(ws, cols):
    for col, text in enumerate(cols, 1):
        c = ws.cell(row=1, column=col, value=text)
        c.font = HEADER_FONT
        c.fill = HEADER_FILL
        c.alignment = Alignment(horizontal="center")

def _auto_width(ws):
    for col in ws.columns:
        max_len = max((len(str(c.value)) if c.value else 0 for c in col), default=8)
        ws.column_dimensions[col[0].column_letter].width = max(max_len + 2, 12)

# AADT values [veh/day]
AADT = {
    'Evergreen Rd Southbound': 2518,
    'Evergreen Rd Northbound': 3042,
    'Hubbard Rd Eastbound':    7280,
    'Hubbard Rd Westbound':    4497,
}

# Raw hourly vehicle counts (not yet normalised)
RAW = {
    'Evergreen Rd Southbound': [
        14,  1,  7,  8,  9, 38, 66,  75, 134, 132, 152, 169,
       172, 185, 206, 203, 206, 212, 144, 126, 100,  57,  39, 16,
    ],
    'Evergreen Rd Northbound': [
        23, 10, 10,  5, 14, 21,  85, 109, 155, 155, 173, 165,
       229, 223, 275, 262, 237, 249, 198, 138, 127,  78,  46, 35,
    ],
    'Hubbard Rd Eastbound': [
        36,  38,  17,  22,  17,  38,  72, 108, 161, 238, 331, 491,
       610, 524, 638, 647, 683, 596, 598, 532, 309, 172, 133,  69,
    ],
    'Hubbard Rd Westbound': [
        31,  18,  16,  19,  21,  51, 136, 192, 236, 236, 330, 335,
       376, 372, 370, 370, 363, 355, 298, 173, 150,  90,  80,  38,
    ],
}

wb = Workbook()
wb.remove(wb.active)  # remove default empty sheet

for road_name, aadt in AADT.items():
    ws = wb.create_sheet(title=road_name)
    _header(ws, ["Hour_1to24", "AADT_veh_per_day", "HourlyCount_raw"])
    raw = RAW[road_name]
    for h, count in enumerate(raw, 1):
        ws.cell(row=h + 1, column=1, value=h)
        ws.cell(row=h + 1, column=2, value=aadt)
        ws.cell(row=h + 1, column=3, value=count)
    _auto_width(ws)

out = PROJECT_DIR / "truth_data.xlsx"
wb.save(out)
print(f"Written: {out}")
