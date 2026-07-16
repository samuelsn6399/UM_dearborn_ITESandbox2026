"""
generate_scenario_config.py
===========================
One-time script: writes projects/UM_Dearborn/scenario_config.xlsx with
three sample scenarios for the UM-Dearborn reference/demo case.

Scenarios included:
  Baseline         — no overrides (reference run)
  SignalRetiming_A — longer green on Evergreen SB to reduce peak-hour queue
  LandUseCampus+   — 50% increase in campus employment (development scenario)
  ModeSplit_Transit — 20% of campus and student housing trips shift to transit

Run once from the repo root:
    python generate_scenario_config.py
"""

from pathlib import Path
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment

PROJECT_DIR = Path("projects/UM_Dearborn")

HEADER_FONT = Font(bold=True, color="FFFFFF")
HEADER_FILL = PatternFill("solid", fgColor="2E4057")
CENTER      = Alignment(horizontal="center")


def _header(ws, cols):
    for col, text in enumerate(cols, 1):
        c = ws.cell(row=1, column=col, value=text)
        c.font      = HEADER_FONT
        c.fill      = HEADER_FILL
        c.alignment = CENTER


def _auto_width(ws):
    for col in ws.columns:
        max_len = max((len(str(c.value)) if c.value else 0 for c in col), default=8)
        ws.column_dimensions[col[0].column_letter].width = max(max_len + 2, 14)


wb = Workbook()

# ── ScenarioList ─────────────────────────────────────────────────────────────
ws = wb.active
ws.title = "ScenarioList"
_header(ws, ["ScenarioName", "BaseProject", "Description"])
scenarios = [
    ("Baseline",          "UM_Dearborn", "No overrides — reference run"),
    ("SignalRetiming_A",  "UM_Dearborn", "Longer green on Evergreen SB to reduce AM queue"),
    ("LandUseCampus+",    "UM_Dearborn", "50% increase in MainCampus employment"),
    ("ModeSplit_Transit",  "UM_Dearborn", "20% transit mode shift for Campus and StudentHousing"),
]
for r, row in enumerate(scenarios, 2):
    for c, val in enumerate(row, 1):
        ws.cell(row=r, column=c, value=val)
_auto_width(ws)

# ── SignalOverrides ───────────────────────────────────────────────────────────
ws2 = wb.create_sheet("SignalOverrides")
_header(ws2, ["ScenarioName", "RoadName", "Green_s", "Red_s",
              "Qsat_per_lane_vehsperlane"])
signal_rows = [
    # SignalRetiming_A: extend green by 15 s on SB, reduce red accordingly
    ("SignalRetiming_A", "Evergreen Rd Southbound", 60, 60, round(1900/3600, 6)),
]
for r, row in enumerate(signal_rows, 2):
    for c, val in enumerate(row, 1):
        ws2.cell(row=r, column=c, value=val)
_auto_width(ws2)

# ── LandUseOverrides ─────────────────────────────────────────────────────────
ws3 = wb.create_sheet("LandUseOverrides")
_header(ws3, ["ScenarioName", "ZoneName", "Employment", "Enrollment",
              "RetailArea_sqft"])
# Baseline MainCampus: ~2000 jobs, ~8000 students, 0 retail (from TripRateData)
# LandUseCampus+: +50% employment
landuse_rows = [
    ("LandUseCampus+", "MainCampus", 3000, 8000, 0),
]
for r, row in enumerate(landuse_rows, 2):
    for c, val in enumerate(row, 1):
        ws3.cell(row=r, column=c, value=val)
_auto_width(ws3)

# ── ModeSplitOverrides ────────────────────────────────────────────────────────
ws4 = wb.create_sheet("ModeSplitOverrides")
_header(ws4, ["ScenarioName", "ZoneName", "AutoShare", "TransitShare",
              "BikeShare", "WalkShare"])
# Baseline auto_share = 1/1.25 = 0.80 (80% auto trips implied by occupancy)
# ModeSplit_Transit: 20% reduction in auto for campus and student housing
modesplit_rows = [
    ("ModeSplit_Transit", "MainCampus",      0.64, 0.20, 0.08, 0.08),
    ("ModeSplit_Transit", "StudentHousing",  0.64, 0.20, 0.08, 0.08),
]
for r, row in enumerate(modesplit_rows, 2):
    for c, val in enumerate(row, 1):
        ws4.cell(row=r, column=c, value=val)
_auto_width(ws4)

# ── QuickTuneOverrides ────────────────────────────────────────────────────────
ws5 = wb.create_sheet("QuickTuneOverrides")
_header(ws5, ["ScenarioName", "Key", "ScaleFactor"])
# Example: no QuickTune changes for demo scenarios
# (rows can be added here for scenario-specific calibration)
_auto_width(ws5)

out = PROJECT_DIR / "scenario_config.xlsx"
wb.save(out)
print(f"Written: {out}")
