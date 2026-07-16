"""
generate_project_configs.py
===========================
One-time script: writes the UM-Dearborn project configuration xlsx files
  - projects/UM_Dearborn/corridor_config.xlsx
  - projects/UM_Dearborn/taz_config.xlsx

Run once from the repo root:
    python generate_project_configs.py
"""

import openpyxl
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from pathlib import Path


PROJECT_DIR = Path("projects/UM_Dearborn")

# ─── style helpers ────────────────────────────────────────────────────────────
HEADER_FONT  = Font(bold=True, color="FFFFFF")
HEADER_FILL  = PatternFill("solid", fgColor="2E4057")
CENTER       = Alignment(horizontal="center")

def _header(ws, row, cols):
    for col, text in enumerate(cols, 1):
        c = ws.cell(row=row, column=col, value=text)
        c.font  = HEADER_FONT
        c.fill  = HEADER_FILL
        c.alignment = CENTER

def _auto_width(ws):
    for col in ws.columns:
        max_len = max((len(str(c.value)) if c.value else 0 for c in col), default=8)
        ws.column_dimensions[col[0].column_letter].width = max(max_len + 2, 10)


# ═══════════════════════════════════════════════════════════════════════════════
#  corridor_config.xlsx
# ═══════════════════════════════════════════════════════════════════════════════
def write_corridor_config():
    wb = Workbook()

    # ── Sheet 1: Corridors ──────────────────────────────────────────────────
    ws = wb.active
    ws.title = "Corridors"
    _header(ws, 1, ["Name", "Idx", "Length_ft", "BoundaryIdx_In", "BoundaryIdx_Out"])
    corridors = [
        ("Evergreen Rd Southbound", 1, 6500, 4, 5),
        ("Evergreen Rd Northbound", 2, 6500, 5, 4),
        ("Hubbard Rd Eastbound",    3, 4500, 0, 6),
        ("Hubbard Rd Westbound",    4, 4500, 6, 0),
    ]
    for r, row in enumerate(corridors, 2):
        for c, val in enumerate(row, 1):
            ws.cell(row=r, column=c, value=val)
    _auto_width(ws)

    # ── Sheet 2: LaneSegments ───────────────────────────────────────────────
    ws2 = wb.create_sheet("LaneSegments")
    _header(ws2, 1, ["CorridorName", "XStart_ft", "XEnd_ft", "NLanes"])
    lane_rows = [
        # Evergreen SB
        ("Evergreen Rd Southbound",    1, 2000, 4),
        ("Evergreen Rd Southbound", 2001, 3000, 3),
        ("Evergreen Rd Southbound", 3001, 3500, 5),
        ("Evergreen Rd Southbound", 3501, 4500, 3),
        ("Evergreen Rd Southbound", 4501, 5500, 2),
        ("Evergreen Rd Southbound", 5501, 6500, 3),
        # Evergreen NB
        ("Evergreen Rd Northbound",    1,  500, 3),
        ("Evergreen Rd Northbound",  501, 2500, 2),
        ("Evergreen Rd Northbound", 2501, 6500, 3),
        # Hubbard EB
        ("Hubbard Rd Eastbound",       1, 1000, 2),
        ("Hubbard Rd Eastbound",    1001, 1500, 3),
        ("Hubbard Rd Eastbound",    1501, 3500, 2),
        ("Hubbard Rd Eastbound",    3501, 4500, 3),
        # Hubbard WB
        ("Hubbard Rd Westbound",       1, 2000, 3),
        ("Hubbard Rd Westbound",    2001, 3000, 2),
        ("Hubbard Rd Westbound",    3001, 3500, 3),
        ("Hubbard Rd Westbound",    3501, 4500, 2),
    ]
    for r, row in enumerate(lane_rows, 2):
        for c, val in enumerate(row, 1):
            ws2.cell(row=r, column=c, value=val)
    _auto_width(ws2)

    # ── Sheet 3: SpeedSegments ──────────────────────────────────────────────
    ws3 = wb.create_sheet("SpeedSegments")
    _header(ws3, 1, ["CorridorName", "XStart_ft", "XEnd_ft", "Speed_mph"])
    speed_rows = [
        # Evergreen SB — 40 mph except 30 mph zone
        ("Evergreen Rd Southbound",    1, 3500, 40),
        ("Evergreen Rd Southbound", 3501, 5500, 30),
        ("Evergreen Rd Southbound", 5501, 6500, 40),
        # Evergreen NB — uniform 40 mph
        ("Evergreen Rd Northbound",    1, 6500, 40),
        # Hubbard EB — uniform 45 mph
        ("Hubbard Rd Eastbound",       1, 4500, 45),
        # Hubbard WB — uniform 45 mph
        ("Hubbard Rd Westbound",       1, 4500, 45),
    ]
    for r, row in enumerate(speed_rows, 2):
        for c, val in enumerate(row, 1):
            ws3.cell(row=r, column=c, value=val)
    _auto_width(ws3)

    # ── Sheet 4: Signals ────────────────────────────────────────────────────
    ws4 = wb.create_sheet("Signals")
    _header(ws4, 1, ["CorridorName", "SignalX_ft", "Green_s", "Red_s", "Qsat_per_lane_vehsperlane"])
    signal_rows = [
        ("Evergreen Rd Southbound", 6000, 45, 75, round(1900/3600, 6)),
        ("Evergreen Rd Northbound",  500, 45, 75, round(1900/3600, 6)),
        ("Hubbard Rd Eastbound",    4200, 45, 75, round(1900/3600, 6)),
        ("Hubbard Rd Westbound",     500, 45, 75, round(1900/3600, 6)),
    ]
    for r, row in enumerate(signal_rows, 2):
        for c, val in enumerate(row, 1):
            ws4.cell(row=r, column=c, value=val)
    _auto_width(ws4)

    out = PROJECT_DIR / "corridor_config.xlsx"
    wb.save(out)
    print(f"Written: {out}")


# ═══════════════════════════════════════════════════════════════════════════════
#  taz_config.xlsx
# ═══════════════════════════════════════════════════════════════════════════════
def write_taz_config():
    wb = Workbook()

    # ── Sheet 1: Zones ──────────────────────────────────────────────────────
    ws = wb.active
    ws.title = "Zones"
    _header(ws, 1, [
        "ZoneName", "xLocation_ft", "yLocation_ft",
        "PeakArrive_hr", "SigmaArrive_hr",
        "PeakDepart_hr", "SigmaDepart_hr",
    ])
    zones = [
        ("MainCampus",      3500,      0, 14, 4, 14, 4),
        ("ShoppingCenter",  6000,      0, 14, 4, 14, 4),
        ("StudentHousing",  2000,   1900, 14, 4, 14, 4),
        ("NorthBoundary", -10000,      0, 16, 4, 16, 4),
        ("SouthBoundary",  16500,      0, 16, 4, 16, 4),
        ("EastBoundary",    2000,  14500, 14, 5, 14, 5),
    ]
    for r, row in enumerate(zones, 2):
        for c, val in enumerate(row, 1):
            ws.cell(row=r, column=c, value=val)
    _auto_width(ws)

    # ── Sheet 2: AccessPoints ───────────────────────────────────────────────
    ws2 = wb.create_sheet("AccessPoints")
    _header(ws2, 1, ["TazIndex", "RoadName", "XLocal_ft", "Split", "AccessPointName"])
    ap_rows = [
        (1, "Evergreen Rd Southbound", 1700, 0.10, "University Secondary Entrance 1"),
        (1, "Evergreen Rd Southbound", 3200, 0.60, "University Primary Entrance"),
        (1, "Evergreen Rd Southbound", 4500, 0.20, "University Tertiary Entrance"),
        (1, "Evergreen Rd Southbound", 5400, 0.10, "University Secondary Entrance 2"),
        (1, "Evergreen Rd Northbound", 1100, 0.10, "University Secondary Entrance 1"),
        (1, "Evergreen Rd Northbound", 2000, 0.20, "University Primary Entrance"),
        (1, "Evergreen Rd Northbound", 3300, 0.60, "University Tertiary Entrance"),
        (1, "Evergreen Rd Northbound", 4800, 0.10, "University Secondary Entrance 2"),
        (2, "Evergreen Rd Southbound", 6000, 1.00, "Shopping Center"),
        (2, "Evergreen Rd Northbound",  500, 1.00, "Shopping Center"),
        (3, "Hubbard Rd Eastbound",    1200, 0.75, "Student Housing"),
        (3, "Hubbard Rd Eastbound",    3000, 0.25, "Student Housing"),
        (3, "Hubbard Rd Westbound",    1500, 0.25, "Student Housing"),
        (3, "Hubbard Rd Westbound",    3300, 0.75, "Student Housing"),
    ]
    for r, row in enumerate(ap_rows, 2):
        for c, val in enumerate(row, 1):
            ws2.cell(row=r, column=c, value=val)
    _auto_width(ws2)

    # ── Sheet 3: Intersections ──────────────────────────────────────────────
    ws3 = wb.create_sheet("Intersections")
    _header(ws3, 1, ["RoadName", "XLocal_ft", "ExternalTazIndices"])
    intr_rows = [
        ("Evergreen Rd Southbound", 2100, "3,6"),
        ("Evergreen Rd Northbound", 4400, "3,6"),
        ("Hubbard Rd Eastbound",       0, "1,4,5"),
        ("Hubbard Rd Westbound",    4400, "1,4,5"),
    ]
    for r, row in enumerate(intr_rows, 2):
        for c, val in enumerate(row, 1):
            ws3.cell(row=r, column=c, value=val)
    _auto_width(ws3)

    # ── Sheet 4: ODAccess_Depart ────────────────────────────────────────────
    zone_names = ["MainCampus","ShoppingCenter","StudentHousing",
                  "NorthBoundary","SouthBoundary","EastBoundary"]
    road_names = ["Evergreen Rd Southbound","Evergreen Rd Northbound",
                  "Hubbard Rd Eastbound","Hubbard Rd Westbound"]

    ev_sb_dep = [
        [0, 1, 0, 0, 1, 0],
        [0, 0, 0, 0, 1, 0],
        [1, 1, 0, 0, 1, 0],
        [1, 1, 1, 0, 1, 1],
        [0, 0, 0, 0, 0, 0],
        [1, 1, 0, 0, 1, 0],
    ]
    ev_nb_dep = [[ev_sb_dep[j][i] for j in range(6)] for i in range(6)]   # transpose
    hub_eb_dep = [
        [0, 0, 1, 0, 0, 1],
        [0, 0, 1, 0, 0, 1],
        [0, 0, 0, 0, 0, 1],
        [0, 0, 1, 0, 0, 1],
        [0, 0, 1, 0, 0, 1],
        [0, 0, 0, 0, 0, 0],
    ]
    hub_wb_dep = [[hub_eb_dep[j][i] for j in range(6)] for i in range(6)]

    for sheet_name, matrix in [("ODAccess_Depart", ev_sb_dep),
                                 ("ODAccess_Depart_NB", ev_nb_dep),
                                 ("ODAccess_Depart_EB", hub_eb_dep),
                                 ("ODAccess_Depart_WB", hub_wb_dep)]:
        ws_od = wb.create_sheet(sheet_name)
        # First row: road label + zone names
        ws_od.cell(row=1, column=1, value="Origin\\Destination")
        for c, zn in enumerate(zone_names, 2):
            ws_od.cell(row=1, column=c, value=zn)
        ws_od.cell(row=1, column=1).font = HEADER_FONT
        ws_od.cell(row=1, column=1).fill = HEADER_FILL
        for c in range(2, 8):
            ws_od.cell(row=1, column=c).font = HEADER_FONT
            ws_od.cell(row=1, column=c).fill = HEADER_FILL
        for r, (rname, row) in enumerate(zip(zone_names, matrix), 2):
            ws_od.cell(row=r, column=1, value=rname)
            for c, val in enumerate(row, 2):
                ws_od.cell(row=r, column=c, value=val)
        _auto_width(ws_od)

    # ── Sheet: QuickTune ────────────────────────────────────────────────────
    ws_qt = wb.create_sheet("QuickTune")
    _header(ws_qt, 1, ["Key", "ScaleFactor", "Description"])
    qt_rows = [
        ("SB_in",  1.0, "NorthBoundary departures onto Evergreen SB"),
        ("SB_out", 1.0, "SouthBoundary arrivals from Evergreen SB"),
        ("NB_in",  1.2, "SouthBoundary departures onto Evergreen NB"),
        ("NB_out", 1.2, "NorthBoundary arrivals from Evergreen NB"),
        ("EB_in",  2.5, "Intersection departures onto Hubbard EB"),
        ("EB_out", 2.1, "EastBoundary arrivals from Hubbard EB"),
        ("WB_in",  1.3, "EastBoundary departures onto Hubbard WB"),
        ("WB_out", 1.3, "Intersection arrivals from Hubbard WB"),
    ]
    for r, row in enumerate(qt_rows, 2):
        for c, val in enumerate(row, 1):
            ws_qt.cell(row=r, column=c, value=val)
    _auto_width(ws_qt)

    out = PROJECT_DIR / "taz_config.xlsx"
    wb.save(out)
    print(f"Written: {out}")


# ── Main ───────────────────────────────────────────────────────────────────────
if __name__ == "__main__":
    write_corridor_config()
    write_taz_config()
    print("Done.")
