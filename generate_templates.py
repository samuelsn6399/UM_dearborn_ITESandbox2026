"""
generate_templates.py
=====================
One-time script: creates blank Excel template files in templates/ with:
  - Column headers
  - Data-type annotations in row 2 (grey, italic)
  - One example row per sheet

Run once from the repo root:
    python generate_templates.py
"""

from pathlib import Path
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

TEMPLATES_DIR = Path('templates')
TEMPLATES_DIR.mkdir(exist_ok=True)

HDR_FILL  = PatternFill('solid', fgColor='3b5998')
HDR_FONT  = Font(bold=True, color='FFFFFF', size=11)
ANN_FONT  = Font(italic=True, color='888888', size=9)
ANN_FILL  = PatternFill('solid', fgColor='F7F8FA')
EX_FILL   = PatternFill('solid', fgColor='EAF4FB')
EX_FONT   = Font(color='1a1a2e', size=10)

def _write_sheet(ws, headers, annotations, examples):
    """Write headers (row 1), annotations (row 2), examples (rows 3+)."""
    # Headers
    for c, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=c, value=h)
        cell.font = HDR_FONT
        cell.fill = HDR_FILL
        cell.alignment = Alignment(horizontal='center')
    # Annotations
    for c, a in enumerate(annotations, 1):
        cell = ws.cell(row=2, column=c, value=a)
        cell.font = ANN_FONT
        cell.fill = ANN_FILL
        cell.alignment = Alignment(wrap_text=True)
    # Examples
    for r, row_data in enumerate(examples, 3):
        for c, v in enumerate(row_data, 1):
            cell = ws.cell(row=r, column=c, value=v)
            cell.font = EX_FONT
            cell.fill = EX_FILL
    # Column widths
    for c in range(1, len(headers) + 1):
        ws.column_dimensions[get_column_letter(c)].width = 22
    # Freeze header rows
    ws.freeze_panes = 'A3'


# ── corridor_config_template.xlsx ─────────────────────────────────────────────
wb = openpyxl.Workbook()

ws = wb.active; ws.title = 'Corridors'
_write_sheet(ws,
    headers=['Name', 'Idx', 'Length_ft', 'BoundaryIdx_In', 'BoundaryIdx_Out'],
    annotations=['text · unique road name',
                 'int · 1-based',
                 'number · feet',
                 'int · TAZ index or 0=intersection',
                 'int · TAZ index or 0=intersection'],
    examples=[['Main St Northbound', 1, 7920, 3, 4],
              ['Main St Southbound', 2, 7920, 4, 3]])

ws2 = wb.create_sheet('LaneSegments')
_write_sheet(ws2,
    headers=['CorridorName', 'XStart_ft', 'XEnd_ft', 'NLanes'],
    annotations=['text · must match Corridors.Name',
                 'number · ft from inflow end',
                 'number · ft',
                 'int · through lanes'],
    examples=[['Main St Northbound', 0, 2640, 3],
              ['Main St Northbound', 2641, 5280, 4],
              ['Main St Northbound', 5281, 7920, 3]])

ws3 = wb.create_sheet('SpeedSegments')
_write_sheet(ws3,
    headers=['CorridorName', 'XStart_ft', 'XEnd_ft', 'Speed_mph'],
    annotations=['text · must match Corridors.Name',
                 'number · ft',
                 'number · ft',
                 'number · posted free-flow speed [mph]'],
    examples=[['Main St Northbound', 0, 7920, 35]])

ws4 = wb.create_sheet('Signals')
_write_sheet(ws4,
    headers=['CorridorName', 'SignalX_ft', 'Green_s', 'Red_s',
             'Qsat_per_lane_vehsperlane'],
    annotations=['text',
                 'number · signal position [ft] from inflow',
                 'number · green phase [s]',
                 'number · red phase [s]',
                 'number · 0.528 = 1900 veh/hr/lane'],
    examples=[['Main St Northbound', 5280, 45, 75, 0.528],
              ['Main St Southbound', 2640, 45, 75, 0.528]])

wb.save(TEMPLATES_DIR / 'corridor_config_template.xlsx')
print('Saved: templates/corridor_config_template.xlsx')


# ── taz_config_template.xlsx ──────────────────────────────────────────────────
wb2 = openpyxl.Workbook()

ws = wb2.active; ws.title = 'Zones'
_write_sheet(ws,
    headers=['ZoneName', 'xLocation_ft', 'yLocation_ft',
             'PeakArrive_hr', 'SigmaArrive_hr',
             'PeakDepart_hr', 'SigmaDepart_hr'],
    annotations=['text · unique; matches HH/TripRate sheet names',
                 'number · ft',
                 'number · ft',
                 'number · peak arrival hour [1-24]',
                 'number · arrival spread [hours]',
                 'number · peak departure hour [1-24]',
                 'number · departure spread [hours]'],
    examples=[['CommercialZone', 2640, 0, 9, 2, 17, 2],
              ['ResidentialZone', 1320, 500, 7, 1.5, 16, 2],
              ['NorthBoundary', -5000, 0, 16, 4, 16, 4]])

ws2 = wb2.create_sheet('AccessPoints')
_write_sheet(ws2,
    headers=['TazIndex', 'RoadName', 'XLocal_ft', 'Split', 'AccessPointName'],
    annotations=['int · 1-based row in Zones sheet',
                 'text · matches corridor_config.xlsx',
                 'number · ft along road',
                 'number · fraction [0-1]; same TAZ+Road rows must sum to 1',
                 'text · label for plots'],
    examples=[[1, 'Main St Northbound', 1320, 1.0, 'North Driveway'],
              [2, 'Main St Northbound', 4400, 0.6, 'Shopping Entrance'],
              [2, 'Main St Northbound', 4800, 0.4, 'Shopping Exit']])

ws3 = wb2.create_sheet('Intersections')
_write_sheet(ws3,
    headers=['RoadName', 'XLocal_ft', 'ExternalTazIndices'],
    annotations=['text · road with boundary_idx=0',
                 'number · ft',
                 'text · comma-separated TAZ indices that feed this intersection'],
    examples=[['Cross St Eastbound', 0, '3,4'],
              ['Cross St Westbound', 7920, '3,4']])

# ODAccess sheets (one per road direction)
for sheet_suffix in ['Depart', 'Depart_NB', 'Depart_EB', 'Depart_WB']:
    ws_od = wb2.create_sheet(f'ODAccess_{sheet_suffix}')
    zones = ['CommercialZone', 'ResidentialZone', 'NorthBoundary']
    # Header row
    ws_od.cell(1, 1, 'Origin\\Destination').font = HDR_FONT
    ws_od.cell(1, 1).fill = HDR_FILL
    for c, z in enumerate(zones, 2):
        ws_od.cell(1, c, z).font = HDR_FONT
        ws_od.cell(1, c).fill = HDR_FILL
    # Zone rows
    example_mat = [[0, 1, 1], [1, 0, 1], [1, 1, 0]]
    for r, (z, row_data) in enumerate(zip(zones, example_mat), 2):
        ws_od.cell(r, 1, z)
        for c, v in enumerate(row_data, 2):
            ws_od.cell(r, c, v).fill = EX_FILL
    ws_od.freeze_panes = 'B2'
    for c in range(1, 5):
        ws_od.column_dimensions[get_column_letter(c)].width = 20

ws_qt = wb2.create_sheet('QuickTune')
_write_sheet(ws_qt,
    headers=['Key', 'ScaleFactor'],
    annotations=['text · pattern: {DirectionAbbrev}_{in|out}',
                 'number · 1.0 = no change'],
    examples=[['NB_in', 1.0], ['NB_out', 1.0],
              ['SB_in', 1.0], ['SB_out', 1.0]])

wb2.save(TEMPLATES_DIR / 'taz_config_template.xlsx')
print('Saved: templates/taz_config_template.xlsx')


# ── truth_data_template.xlsx ──────────────────────────────────────────────────
wb3 = openpyxl.Workbook()

for i, road_name in enumerate(['Road1_Direction', 'Road2_Direction']):
    ws_t = wb3.active if i == 0 else wb3.create_sheet(road_name)
    if i == 0:
        ws_t.title = road_name
    _write_sheet(ws_t,
        headers=['Hour', 'AADT', 'HourlyFraction'],
        annotations=['int · 1-24',
                     'number · Annual Average Daily Traffic [veh/day]',
                     'number · fraction of AADT in this hour; all 24 rows must sum to 1.0'],
        examples=[[h, 5000, round(1/24, 6)] for h in range(1, 25)])

wb3.save(TEMPLATES_DIR / 'truth_data_template.xlsx')
print('Saved: templates/truth_data_template.xlsx')

print('\nAll templates generated successfully.')
